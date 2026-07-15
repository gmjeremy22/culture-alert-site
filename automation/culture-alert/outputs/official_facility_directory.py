import argparse
import csv
import difflib
import json
import math
import re
import sqlite3
from bisect import bisect_right
from collections import Counter
from datetime import datetime
from pathlib import Path
from urllib.parse import urlparse

import openpyxl


BASE_DIR = Path(__file__).resolve().parent
DEFAULT_DB = BASE_DIR / "culture-alert.sqlite"
DEFAULT_WORKBOOK = (
    BASE_DIR.parent / "work" / "official-directory" / "2025-cultural-facilities-directory.xlsx"
)
DEFAULT_CSV = BASE_DIR / "official-facility-directory.csv"
DEFAULT_SEOUL_CSV = BASE_DIR / "seoul-official-institutions.csv"
DEFAULT_METRICS = BASE_DIR / "institution-scale-metrics.csv"
DEFAULT_JSON = BASE_DIR / "official-facility-directory-audit-report.json"
DEFAULT_REPORT = BASE_DIR / "official-facility-directory-audit-report.md"
DEFAULT_HTML = BASE_DIR / "keyword-recommendation-report.html"
SOURCE_URL = (
    "https://www.mcst.go.kr/site/s_policy/dept/deptView.jsp?"
    "pDataCD=0417000000&pSeq=2078&pType="
)
DIRECTORY_YEAR = 2025
REFERENCE_DATE = "2025-01-01"
CAPITAL_REGIONS = {"서울", "경기", "인천"}
SHEET_CATEGORIES = {"박물관": "박물관", "미술관": "미술관"}


NAME_ALIASES = {
    "국립경찰박물관": "경찰박물관",
    "농협농업박물관": "농업박물관",
    "국립현대미술관서울": "국립현대미술관서울관",
    "국립현대미술관과천": "국립현대미술관과천관",
    "국립현대미술관덕수궁": "국립현대미술관덕수궁관",
    "서울시립미술관서소문본관": "서울시립미술관",
    "서울공예박물관": "서울공예박물관",
    "리움미술관": "리움미술관",
    "수도국산달동네박물관휴관": "수도국산달동네박물관",
    "코리아나미술관": "코리아나미술관스페이스씨",
    "한가람디자인미술관": "예술의전당한가람디자인미술관",
    "한가람미술관": "예술의전당한가람미술관",
    "고양시립아람미술관": "고양아람누리아람미술관",
}


def compact(value):
    return " ".join(str(value or "").replace("\u00a0", " ").split())


def number(value):
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return max(0.0, float(value))
    matched = re.search(r"-?[\d,]+(?:\.\d+)?", str(value))
    if not matched:
        return None
    try:
        return max(0.0, float(matched.group(0).replace(",", "")))
    except ValueError:
        return None


def normalize_name(value):
    text = compact(value).casefold()
    text = re.sub(r"\((?:재|사|주|유)\)", "", text)
    text = re.sub(r"재단법인|사단법인|주식회사|유한회사", "", text)
    text = text.replace("ㆍ", "").replace("·", "")
    return re.sub(r"[^0-9a-z가-힣]", "", text)


def alias_key(value):
    normalized = normalize_name(value)
    return NAME_ALIASES.get(normalized, normalized)


def normalize_url(value):
    text = compact(value)
    if not text or text.lower() in {"x", "없음", "해당없음", "-"}:
        return ""
    matched = re.search(r"(?:https?://|www\.)[^\s,;]+", text, re.I)
    if matched:
        text = matched.group(0).rstrip(".)]")
    else:
        text = text.split()[0].strip(",;")
    if text.startswith("www."):
        text = "https://" + text
    elif not re.match(r"https?://", text, re.I) and "." in text:
        text = "https://" + text
    if not re.match(r"https?://", text, re.I):
        return ""
    return text


def url_domain(value):
    url = normalize_url(value)
    if not url:
        return ""
    return urlparse(url).netloc.casefold().removeprefix("www.").split(":", 1)[0]


def percentile(value, population):
    if value is None or not population:
        return None
    transformed = math.log1p(value)
    ordered = sorted(math.log1p(item) for item in population if item is not None)
    if len(ordered) == 1:
        return 1.0
    return (bisect_right(ordered, transformed) - 1) / (len(ordered) - 1)


def governance_score(ownership, registration_type):
    ownership_text = compact(ownership).replace(" ", "")
    registration_text = compact(registration_type).replace(" ", "")
    ownership_points = 10 if "국립" in ownership_text else 6 if "공립" in ownership_text else 4 if "대학" in ownership_text else 2
    registration_points = 5 if "1종" in registration_text else 2 if "2종" in registration_text else 0
    return ownership_points + registration_points


def extract_rows(workbook_path):
    workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    rows = []
    for sheet_name, category in SHEET_CATEGORIES.items():
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"필수 시트가 없습니다: {sheet_name}")
        sheet = workbook[sheet_name]
        current_region = ""
        for row_number, values in enumerate(sheet.iter_rows(min_row=10, values_only=True), start=10):
            region = compact(values[1] if len(values) > 1 else "")
            if region:
                current_region = region
            if current_region not in CAPITAL_REGIONS:
                continue
            source_name = compact(values[5] if len(values) > 5 else "")
            serial = values[0] if values else None
            if not source_name or serial in (None, ""):
                continue
            rows.append(
                {
                    "directory_year": DIRECTORY_YEAR,
                    "reference_date": REFERENCE_DATE,
                    "source_name": source_name,
                    "region": current_region,
                    "city": compact(values[2] if len(values) > 2 else ""),
                    "facility_type": category,
                    "ownership": compact(values[3] if len(values) > 3 else ""),
                    "registration_type": compact(values[4] if len(values) > 4 else ""),
                    "address": compact(values[6] if len(values) > 6 else ""),
                    "phone": compact(values[7] if len(values) > 7 else ""),
                    "opening_date": compact(values[8] if len(values) > 8 else ""),
                    "registration_status": compact(values[9] if len(values) > 9 else ""),
                    "registration_number": compact(values[11] if len(values) > 11 else ""),
                    "official_url": normalize_url(values[12] if len(values) > 12 else ""),
                    "exhibition_area_sqm": number(values[18] if len(values) > 18 else None),
                    "collection_count": number(values[33] if len(values) > 33 else None),
                    "annual_visitors": number(values[46] if len(values) > 46 else None),
                    "source_sheet": sheet_name,
                    "source_row": row_number,
                    "source_url": SOURCE_URL,
                }
            )
    return rows


def score_rows(rows):
    fields = {
        "annual_visitors": 55,
        "collection_count": 20,
        "exhibition_area_sqm": 10,
    }
    populations = {
        field: [row[field] for row in rows if row[field] is not None]
        for field in fields
    }
    for row in rows:
        weighted_total = 0.0
        available_weight = 0.0
        available_metrics = 0
        for field, weight in fields.items():
            rank = percentile(row[field], populations[field])
            if rank is None:
                continue
            weighted_total += rank * weight
            available_weight += weight
            available_metrics += 1
        evidence = weighted_total / available_weight * sum(fields.values()) if available_weight else 0
        row["official_scale_score"] = round(
            min(100.0, evidence + governance_score(row["ownership"], row["registration_type"])),
            1,
        )
        row["data_confidence"] = round(available_metrics / len(fields), 2)
    return rows


def ensure_schema(conn):
    schema_path = BASE_DIR / "culture-alert-schema.sql"
    conn.executescript(schema_path.read_text(encoding="utf-8"))


def institution_rows(conn):
    columns = [description[0] for description in conn.execute("SELECT * FROM institutions LIMIT 0").description]
    return [dict(zip(columns, row)) for row in conn.execute("SELECT * FROM institutions ORDER BY active DESC, id")]


def name_similarity(left, right):
    return difflib.SequenceMatcher(None, normalize_name(left), normalize_name(right)).ratio()


def match_row(source, institutions):
    region_candidates = [row for row in institutions if row.get("region") == source["region"]]
    source_key = alias_key(source["source_name"])
    exact = [row for row in region_candidates if alias_key(row["name"]) == source_key]
    if len(exact) == 1:
        return exact[0], "name_exact", 1.0, exact[0]["name"]
    if len(exact) > 1:
        city_exact = [row for row in exact if compact(row.get("city")) == source["city"]]
        if len(city_exact) == 1:
            return city_exact[0], "name_city_exact", 1.0, city_exact[0]["name"]

    source_domain = url_domain(source["official_url"])
    scored = []
    for institution in region_candidates:
        similarity = name_similarity(source["source_name"], institution["name"])
        same_city = bool(source["city"] and compact(institution.get("city")) == source["city"])
        existing_domains = {
            url_domain(institution.get("exhibition_url")),
            url_domain(institution.get("program_url")),
        } - {""}
        same_domain = bool(source_domain and source_domain in existing_domains)
        score = similarity + (0.22 if same_city else 0) + (0.32 if same_domain else 0)
        scored.append((score, similarity, same_city, same_domain, institution))
    scored.sort(key=lambda item: (item[0], item[4].get("active", 0)), reverse=True)
    if not scored:
        return None, "new", 0.0, ""
    best = scored[0]
    next_score = scored[1][0] if len(scored) > 1 else 0
    domain_match_is_specific = best[3] and best[1] >= 0.90
    if domain_match_is_specific and best[0] - next_score >= 0.05:
        return best[4], "domain_name", round(best[0], 3), best[4]["name"]
    if best[1] >= 0.94 and (best[2] or best[0] - next_score >= 0.08):
        return best[4], "name_fuzzy", round(best[0], 3), best[4]["name"]
    return None, "new", round(best[0], 3), best[4]["name"]


def unique_institution_name(source, institutions):
    used = {row["name"] for row in institutions}
    candidates = [
        source["source_name"],
        f'{source["source_name"]} ({source["city"]})',
        f'{source["source_name"]} ({source["region"]} {source["city"]})',
        f'{source["source_name"]} ({source["source_sheet"]} {source["source_row"]})',
    ]
    return next(name for name in candidates if name not in used)


def priority_for_score(score):
    if score >= 82:
        return 1
    if score >= 70:
        return 2
    return 3


def import_rows(conn, rows, audit_only=False):
    institutions = institution_rows(conn)
    results = []
    for source in rows:
        matched, match_type, match_score, closest_existing_name = match_row(source, institutions)
        if matched:
            institution_id = matched["id"]
            institution_name = matched["name"]
            result_status = "matched"
            if not audit_only:
                preferred_priority = min(int(matched.get("priority") or 3), priority_for_score(source["official_scale_score"]))
                official_url = source["official_url"] if not (matched.get("exhibition_url") or matched.get("program_url")) else matched.get("exhibition_url")
                conn.execute(
                    """
                    UPDATE institutions
                    SET priority=?, category=?, city=COALESCE(NULLIF(city,''), ?),
                        exhibition_url=COALESCE(NULLIF(exhibition_url,''), ?),
                        updated_at=CURRENT_TIMESTAMP
                    WHERE id=?
                    """,
                    (preferred_priority, source["facility_type"], source["city"], official_url, institution_id),
                )
        else:
            institution_name = unique_institution_name(source, institutions)
            institution_id = None
            result_status = "would_insert" if audit_only else "inserted"
            if not audit_only:
                cursor = conn.execute(
                    """
                    INSERT INTO institutions (
                      name, region, city, category, priority, collection_phase,
                      exhibition_url, program_url, notes, active
                    ) VALUES (?, ?, ?, ?, ?, 'directory', ?, NULL, ?, 1)
                    """,
                    (
                        institution_name,
                        source["region"],
                        source["city"],
                        source["facility_type"],
                        priority_for_score(source["official_scale_score"]),
                        source["official_url"] or None,
                        f'{DIRECTORY_YEAR} 전국 문화기반시설 총람 등록 기관. 일정 수집 준비 중.',
                    ),
                )
                institution_id = cursor.lastrowid
                matched = {
                    "id": institution_id,
                    "name": institution_name,
                    "region": source["region"],
                    "city": source["city"],
                    "category": source["facility_type"],
                    "priority": priority_for_score(source["official_scale_score"]),
                    "collection_phase": "directory",
                    "exhibition_url": source["official_url"],
                    "program_url": "",
                    "active": 1,
                }
                institutions.append(matched)

        if not audit_only and institution_id is not None:
            conn.execute(
                """
                INSERT INTO institution_directory_metadata (
                  institution_id, directory_year, reference_date, facility_type,
                  source_name, ownership, registration_type, registration_status,
                  registration_number, address, phone, official_url, opening_date,
                  annual_visitors, collection_count, exhibition_area_sqm,
                  official_scale_score, source_sheet, source_row, source_url,
                  updated_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
                ON CONFLICT(directory_year, source_sheet, source_row) DO UPDATE SET
                  institution_id=excluded.institution_id,
                  reference_date=excluded.reference_date,
                  facility_type=excluded.facility_type,
                  source_name=excluded.source_name,
                  ownership=excluded.ownership,
                  registration_type=excluded.registration_type,
                  registration_status=excluded.registration_status,
                  registration_number=excluded.registration_number,
                  address=excluded.address,
                  phone=excluded.phone,
                  official_url=excluded.official_url,
                  opening_date=excluded.opening_date,
                  annual_visitors=excluded.annual_visitors,
                  collection_count=excluded.collection_count,
                  exhibition_area_sqm=excluded.exhibition_area_sqm,
                  official_scale_score=excluded.official_scale_score,
                  source_url=excluded.source_url,
                  updated_at=CURRENT_TIMESTAMP
                """,
                (
                    institution_id,
                    source["directory_year"],
                    source["reference_date"],
                    source["facility_type"],
                    source["source_name"],
                    source["ownership"],
                    source["registration_type"],
                    source["registration_status"],
                    source["registration_number"],
                    source["address"],
                    source["phone"],
                    source["official_url"],
                    source["opening_date"],
                    source["annual_visitors"],
                    source["collection_count"],
                    source["exhibition_area_sqm"],
                    source["official_scale_score"],
                    source["source_sheet"],
                    source["source_row"],
                    source["source_url"],
                ),
            )
            conn.execute(
                """
                INSERT INTO institution_collection_checks (
                  institution_id, source_name, state, detail, checked_at
                ) VALUES (?, 'official-directory', 'directory_only', ?, CURRENT_TIMESTAMP)
                ON CONFLICT(institution_id, source_name) DO UPDATE SET
                  state=CASE WHEN institution_collection_checks.state IN ('collected','empty','review','failed')
                             THEN institution_collection_checks.state ELSE excluded.state END,
                  detail=CASE WHEN institution_collection_checks.state IN ('collected','empty','review','failed')
                              THEN institution_collection_checks.detail ELSE excluded.detail END,
                  checked_at=CASE WHEN institution_collection_checks.state IN ('collected','empty','review','failed')
                                  THEN institution_collection_checks.checked_at ELSE excluded.checked_at END
                """,
                (institution_id, f'{DIRECTORY_YEAR} 공식 총람 등록 · 일정 수집 준비 중'),
            )

        output = dict(source)
        output.update(
            {
                "match_status": result_status,
                "match_type": match_type,
                "match_score": match_score,
                "closest_existing_name": closest_existing_name,
                "institution_id": institution_id or "",
                "institution_name": institution_name,
            }
        )
        results.append(output)
    return results


def read_exported_rows(path):
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.DictReader(handle))
    for row in rows:
        row["directory_year"] = int(row.get("directory_year") or DIRECTORY_YEAR)
        row["source_row"] = int(float(row.get("source_row") or 0))
        for field in (
            "annual_visitors",
            "collection_count",
            "exhibition_area_sqm",
            "official_scale_score",
            "data_confidence",
        ):
            row[field] = number(row.get(field)) or 0.0
        row["official_url"] = normalize_url(row.get("official_url"))
    return rows


def import_directory_csv(conn, path=DEFAULT_CSV):
    rows = read_exported_rows(path)
    return import_rows(conn, rows, audit_only=False)


def write_csv(rows, path):
    fields = [
        "directory_year", "reference_date", "source_name", "institution_name",
        "region", "city", "facility_type", "ownership", "registration_type",
        "registration_status", "registration_number", "address", "phone",
        "official_url", "opening_date", "annual_visitors", "collection_count",
        "exhibition_area_sqm", "official_scale_score", "data_confidence",
        "match_status", "match_type", "match_score", "institution_id",
        "closest_existing_name",
        "source_sheet", "source_row", "source_url",
    ]
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def write_metrics(rows, path):
    fields = [
        "source_name", "region", "city", "category", "ownership",
        "registration_type", "annual_visitors", "collection_count",
        "exhibition_area_sqm", "official_scale_score", "data_confidence",
        "source_reference_date", "source_sheet", "source_row", "source_url",
    ]
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        for row in sorted(rows, key=lambda item: (-item["official_scale_score"], item["region"], item["source_name"])):
            writer.writerow(
                {
                    "source_name": row["source_name"],
                    "region": row["region"],
                    "city": row["city"],
                    "category": row["facility_type"],
                    "ownership": row["ownership"],
                    "registration_type": row["registration_type"],
                    "annual_visitors": int(row["annual_visitors"] or 0),
                    "collection_count": int(row["collection_count"] or 0),
                    "exhibition_area_sqm": round(row["exhibition_area_sqm"] or 0, 2),
                    "official_scale_score": row["official_scale_score"],
                    "data_confidence": row["data_confidence"],
                    "source_reference_date": row["reference_date"],
                    "source_sheet": row["source_sheet"],
                    "source_row": row["source_row"],
                    "source_url": row["source_url"],
                }
            )


def validate_html(html_path):
    if not html_path.exists():
        return {"exists": False, "complete": False, "error": "html_missing"}
    text = html_path.read_text(encoding="utf-8")
    matched = re.search(r"const institutions = (\[.*?\]);\s*const overlay", text, re.S)
    if not matched:
        return {"exists": True, "complete": False, "error": "institution_payload_missing"}
    try:
        institutions = json.loads(matched.group(1))
    except json.JSONDecodeError as exc:
        return {"exists": True, "complete": False, "error": f"institution_payload_invalid: {exc}"}
    seoul = [row for row in institutions if row.get("region") == "서울"]
    seoul_directory = [row for row in seoul if row.get("directoryEntry")]
    default_recommendation_seoul = 'region: "서울"' in text
    default_directory_seoul = 'let institutionRegion = "서울"' in text
    complete = (
        len(seoul_directory) == 180
        and default_recommendation_seoul
        and default_directory_seoul
    )
    return {
        "exists": True,
        "institutions": len(institutions),
        "seoul_institutions": len(seoul),
        "seoul_directory_institutions": len(seoul_directory),
        "default_recommendation_seoul": default_recommendation_seoul,
        "default_directory_seoul": default_directory_seoul,
        "complete": complete,
    }


def build_summary(rows, db_path, audit_only, html_path):
    by_region = Counter(row["region"] for row in rows)
    by_type = Counter(row["facility_type"] for row in rows)
    by_status = Counter(row["match_status"] for row in rows)
    duplicates = [
        {"name": name, "count": count}
        for name, count in Counter(row["source_name"] for row in rows).items()
        if count > 1
    ]
    with sqlite3.connect(db_path) as conn:
        institution_lookup = {
            row[0]: {"name": row[1], "region": row[2], "category": row[3], "active": row[4]}
            for row in conn.execute(
                "SELECT id, name, region, category, active FROM institutions"
            ).fetchall()
        }
        active_total = conn.execute("SELECT COUNT(*) FROM institutions WHERE active=1").fetchone()[0]
        active_capital = conn.execute(
            "SELECT COUNT(*) FROM institutions WHERE active=1 AND region IN ('서울','경기','인천')"
        ).fetchone()[0]
        metadata_total = conn.execute(
            "SELECT COUNT(*) FROM institution_directory_metadata WHERE directory_year=?",
            (DIRECTORY_YEAR,),
        ).fetchone()[0]
    assignment_counts = Counter(
        int(row["institution_id"])
        for row in rows
        if str(row.get("institution_id") or "").isdigit()
    )
    duplicate_assignments = [
        {
            "institution_id": institution_id,
            "institution_name": institution_lookup.get(institution_id, {}).get("name", ""),
            "official_rows": count,
        }
        for institution_id, count in assignment_counts.items()
        if count > 1
    ]
    validation_errors = []
    for row in rows:
        institution_id = row.get("institution_id")
        if not str(institution_id or "").isdigit():
            validation_errors.append({"source_name": row["source_name"], "error": "institution_id_missing"})
            continue
        institution = institution_lookup.get(int(institution_id))
        if not institution:
            validation_errors.append({"source_name": row["source_name"], "error": "institution_missing"})
            continue
        if institution["region"] != row["region"]:
            validation_errors.append({"source_name": row["source_name"], "error": "region_mismatch"})
        if institution["category"] != row["facility_type"]:
            validation_errors.append({"source_name": row["source_name"], "error": "category_mismatch"})
        if not institution["active"]:
            validation_errors.append({"source_name": row["source_name"], "error": "inactive_institution"})
    seoul_rows = [row for row in rows if row["region"] == "서울"]
    seoul_ids = {
        int(row["institution_id"])
        for row in seoul_rows
        if str(row.get("institution_id") or "").isdigit()
    }
    seoul_errors = [
        error for error in validation_errors
        if any(row["source_name"] == error["source_name"] for row in seoul_rows)
    ]
    return {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "audit_only": audit_only,
        "directory_year": DIRECTORY_YEAR,
        "reference_date": REFERENCE_DATE,
        "source_url": SOURCE_URL,
        "official_rows": len(rows),
        "by_region": dict(sorted(by_region.items())),
        "by_facility_type": dict(sorted(by_type.items())),
        "match_status": dict(sorted(by_status.items())),
        "duplicate_source_names": duplicates,
        "duplicate_institution_assignments": duplicate_assignments,
        "validation_errors": validation_errors,
        "seoul": {
            "official_rows": len(seoul_rows),
            "museum_rows": sum(row["facility_type"] == "박물관" for row in seoul_rows),
            "art_museum_rows": sum(row["facility_type"] == "미술관" for row in seoul_rows),
            "distinct_institution_ids": len(seoul_ids),
            "validation_errors": seoul_errors,
            "complete": len(seoul_rows) == len(seoul_ids) and not seoul_errors,
        },
        "active_institutions_total": active_total,
        "active_capital_institutions": active_capital,
        "metadata_rows": metadata_total,
        "html_validation": validate_html(html_path),
    }


def write_report(summary, path):
    lines = [
        "# 공식 문화기반시설 총람 기관 대조 보고서",
        "",
        f"- 생성 시각: {summary['generated_at']}",
        f"- 원자료: {summary['directory_year']} 전국 문화기반시설 총람 (기준일 {summary['reference_date']})",
        f"- 수도권 박물관·미술관 공식 행: {summary['official_rows']}곳",
        f"- 지역별: {summary['by_region']}",
        f"- 유형별: {summary['by_facility_type']}",
        f"- 대조 결과: {summary['match_status']}",
        f"- 서울 완전성: {summary['seoul']}",
        f"- 중복 기관 연결: {len(summary['duplicate_institution_assignments'])}건",
        f"- 검증 오류: {len(summary['validation_errors'])}건",
        f"- DB 공식 총람 메타데이터: {summary['metadata_rows']}행",
        f"- 현재 활성 수도권 기관: {summary['active_capital_institutions']}곳",
        f"- 생성 HTML 검증: {summary['html_validation']}",
        "",
        "총람 등록과 일정 자동 수집은 별도 상태입니다. 일정이 아직 연결되지 않은 기관도 기관 둘러보기에 유지됩니다.",
    ]
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main():
    parser = argparse.ArgumentParser(description="Import every Seoul/Gyeonggi/Incheon museum and art museum from the official directory.")
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--db", type=Path, default=DEFAULT_DB)
    parser.add_argument("--csv", type=Path, default=DEFAULT_CSV)
    parser.add_argument("--seoul-csv", type=Path, default=DEFAULT_SEOUL_CSV)
    parser.add_argument("--metrics", type=Path, default=DEFAULT_METRICS)
    parser.add_argument("--json", type=Path, default=DEFAULT_JSON)
    parser.add_argument("--report", type=Path, default=DEFAULT_REPORT)
    parser.add_argument("--html", type=Path, default=DEFAULT_HTML)
    parser.add_argument("--audit-only", action="store_true")
    parser.add_argument(
        "--from-csv",
        action="store_true",
        help="Use the committed directory CSV instead of the annual source workbook.",
    )
    args = parser.parse_args()

    rows = read_exported_rows(args.csv) if args.from_csv else score_rows(extract_rows(args.workbook))
    with sqlite3.connect(args.db) as conn:
        ensure_schema(conn)
        results = import_rows(conn, rows, audit_only=args.audit_only)
        if args.audit_only:
            conn.rollback()
        else:
            conn.commit()
    write_csv(results, args.csv)
    write_csv([row for row in results if row["region"] == "서울"], args.seoul_csv)
    write_metrics(results, args.metrics)
    summary = build_summary(results, args.db, args.audit_only, args.html)
    args.json.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    write_report(summary, args.report)
    print(json.dumps(summary, ensure_ascii=True, indent=2))


if __name__ == "__main__":
    main()
